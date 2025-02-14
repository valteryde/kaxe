
try:
    from openpyxl import load_workbook
except ImportError:
    pass

LETTERS = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"


def loadExcel(fname, sheet:str, top:tuple, bottom:tuple, flip:bool=False):
    """
    Load data from an Excel sheet within a specified range.

    Parameters
    ----------
    fname : str
        The filename of the Excel workbook.
    sheet : str
        The name of the sheet to load data from.
    top : tuple
        A tuple (column, row) representing the top-left corner of the range.
    bottom : tuple
        A tuple (column, row) representing the bottom-right corner of the range. 
        If the row is -1, it will be set to the maximum row of the sheet.
    flip : bool, optional
        If True, the data will be transposed (flipped). Defaults to False.

    Returns
    -------
    list
        A 2D list containing the data from the specified range in the Excel sheet.
    """
    
    wb = load_workbook(filename = fname)
    ws = wb[sheet]

    if bottom[1] == -1:
        bottom = [bottom[0],ws.max_row]

    data = []
    for i in range(top[1], bottom[1]+1):
        data.append([])

        for j in range(top[0], bottom[0]+1):
            data[-1].append(ws.cell(row=i, column=j).value)

    if flip:
        data = [[data[i][j] for i in range(len(data))] for j in range(len(data[0]))]

    return data
    
